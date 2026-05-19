# -*- coding: utf-8 -*-
"""
Монгол хэл дээрх хичээлийн транскрипцийн үг, өгүүлбэрийн алдааг засах скрипт.
2 файлын бүх sheet дээрх бичвэрийн алдааг олж засна.
"""
import openpyxl
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# Нийтлэг засварууд (хоёр файлд хамаарах)
# ============================================================
common_replacements = [
    # Давхар зай -> нэг зай
    # (handled separately with regex)
]

# ============================================================
# FILE 1: 50_analysis (2) (1).xlsx - Бүх sheet-ийн засварууд
# ============================================================
file1_replacements = [
    # === Хичээл 48 (Sheet) ===
    # Row 14: хариулаарй -> хариулаарай
    ("хариулаарй", "хариулаарай"),
    # Row 16: биднь -> бид нь
    ("биднь", "бид нь"),
    # Row 24: храиу -> хариу
    ("храиу", "хариу"),
    # Row 24: байхгү -> байхгүй
    ("байхгү юу", "байхгүй юу"),
    # Row 24: гаргсан -> гаргасан
    ("гаргсан", "гаргасан"),
    # Row 26: Зөа -> За
    ("Зөа ", "За "),
    # Row 26: хагса нь -> хагас нь
    ("хагса нь", "хагас нь"),
    # Row 26: хагсатай -> хагастай
    ("хагсатай", "хагастай"),
    # Row 26: хагстай -> хагастай
    ("хагстай", "хагастай"),
    # Row 26: 2 лахь -> 2 дахь
    ("2 лахь", "2 дахь"),
    # Row 28: өөрсддөө -> өөрсдөө
    ("өөрсддөө", "өөрсдөө"),
    # Row 30: C column - "Багшуваасан байна. " -> should be "Багш" (handled specially)
    # Row 30: тарааэ өгьеөө -> тарааж өгъё
    ("тарааэ өгьеөө", "тарааж өгъё"),
    # Row 30: анхааралаа ахндуул -> анхааралаа хандуул
    ("ахндуул", "хандуул"),
    # Row 34: З а тэгвэл -> За тэгвэл
    ("З а тэгвэл", "За тэгвэл"),
    # Row 34: таанй -> танай
    ("таанй", "танай"),
    # Row 34: ёажиглаж -> ажиглаж
    ("ёажиглаж", "ажиглаж"),
    # Row 34: Үа хариугаа -> За хариугаа
    ("Үа хариугаа", "За хариугаа"),
    # Row 34: эн баг -> энэ баг
    ("эн баг", "энэ баг"),
    # Row 34: анхааралаа -> анхааралаа (ok but check)
    # Row 40: зүйт огтолыг -> зүй тогтолыг
    ("зүйт огтолыг", "зүй тогтолыг"),
    # Row 46: эн эв -> энэ эв
    ("эн эв", "энэ эв"),
    # Row 47: өргөж байна -> өсөж байна (context: "2 дахин өргөж")
    ("2 дахин өргөж байна", "2 дахин өсөж байна"),
    # Row 49: Өргөж байна -> Өсөж байна
    ("Өргөж байна.", "Өсөж байна."),
    # Row 50: өшөө -> өөр
    ("өшөө ", "өөр "),
    # Row 58: үржшшлээд -> үржүүлээд
    ("үржшшлээд", "үржүүлээд"),
    # Row 63: 1-"р -> 1-р
    ('1-"р', "1-р"),
    # Row 72: давтрээ -> дэвтрээ
    ("давтрээ", "дэвтрээ"),
    # Row 72: нээгэрэй -> нээгээрэй
    ("нээгэрэй", "нээгээрэй"),
    # Row 78: Задахь -> За дахь
    ("Задахь", "За дахь"),
    # Row 100: болнооо -> болно
    ("болнооо", "болно"),
    # Row 102: хийхиий -> хийхийг
    ("хийхиий", "хийхийг"),
    # Row 104: тэнцүбутархайнуудыг -> тэнцүү бутархайнуудыг
    ("тэнцүбутархайнуудыг", "тэнцүү бутархайнуудыг"),
    # Row 106: ажиллнаа -> ажиллана
    ("ажиллнаа", "ажиллана"),
    # Row 110: Шсвэл -> За тэгвэл (contextually this seems like a typo for "За тэгвэл")
    ("Шсвэл", "За тэгвэл"),
    # Row 112: ажигалаарай -> ажиглаарай
    ("ажигалаарай", "ажиглаарай"),
    # Row 126: 20за -> 20 за
    ("20за ", "20 за "),
    # Row 128: хүүдүүдээ -> хүүхдүүдээ
    ("хүүдүүдээ", "хүүхдүүдээ"),
    # Row 130: гэрэх нь -> гарах нь
    ("гэрэх нь", "гарах нь"),
    # Row 130: гараах -> гарах
    ("гараах ", "гарах "),
    # Row 132: цоихно -> цохино
    ("цоихно", "цохино"),
    # Row 132: лбагийн -> багийн
    ("лбагийн", "багийн"),
    # Row 132: цохиорой -> цохиорой (OK)
    # Row 139: цохииру -> цохиурыг
    ("цохииру ", "цохиурыг "),
    # Row 139: эн ялааны -> энэ ялааны
    ("эн ялааны", "энэ ялааны"),
    # Row 151: гсэн -> гэсэн
    ("гсэн ", "гэсэн "),
    # Row 153: баярллаа -> баярлалаа
    ("баярллаа", "баярлалаа"),
    # Row 153: тавиарай -> тавьаарай
    ("тавиарай", "тавьаарай"),
    
    # === Other sheets found errors ===
    # хичээл 45: Цлиндр -> Цилиндр (D1 subject)
    ("Цлиндр", "Цилиндр"),
    # Хичээл 34, Row 105: хажжут -> хажуут
    ("хажжут", "хажуут"),
    # Хичээл 47: хүүхэл -> хүүхэд (Row 11)
    ("хүүхэл бүр", "хүүхэд бүр"),
    # Хичээл 24: extra spaces in subject
    # Хичээл 20: extra space before name
    # Row 14 in Хичээл 48: малгайг нт -> малгайг нь
    ("малгайг нт", "малгайг нь"),
    
    # Various common typos found across multiple sheets
    ("дэлгэцнээс", "дэлгэцнээс"),  # OK
    ("шүүхамгийн", "шүү хамгийн"),
    ("дарна шүүхамгийн", "дарна шүү. Хамгийн"),
    ("өргөөрөэй", "өргөөрэй"),
    ("хүүхдүүдэ.", "хүүхдүүдээ."),
]

# ============================================================
# FILE 2: DATA_time (1) (1) (1).xlsx - Засварууд
# ============================================================
file2_replacements = [
    # Row 69: B69 end time 25:04m should be after start 25:45m -> 26:04m
    ("25:04m", "26:04m"),  # Only in B69 context
    # Row 77: "27: 26m" -> "27:26m" (space in time)
    ("27: 26m", "27:26m"),
    # Row 97: жиссэн -> жишсэн
    ("жиссэн", "жишсэн"),
    # Row 100: жиссэн -> жишсэн (already covered above)
    # Row 117: шиэдгч -> шидэгч
    ("шиэдгч", "шидэгч"),
    # Row 141: хашир аргаар -> хялбар аргаар
    ("хашир аргаар", "хялбар аргаар"),
]


def fix_double_spaces(text):
    """Давхар зайг нэг зай болгоно (гэхдээ зориудаар хэд хэдэн зай хийсэн газрыг үлдээнэ)."""
    # Replace double+ spaces with single, but preserve intentional formatting
    return re.sub(r'  +', ' ', text)


def apply_fixes(filepath, replacements, output_path, fix_spaces=True):
    """Apply text replacements to an Excel file."""
    wb = openpyxl.load_workbook(filepath)
    changes_made = 0
    change_log = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                        original = cell.value
                        new_value = original

                        # Apply text replacements
                        for wrong, correct in replacements:
                            if wrong != correct and wrong in new_value:
                                new_value = new_value.replace(wrong, correct)

                        # Fix double spaces (optional)
                        if fix_spaces and '  ' in new_value:
                            new_value = fix_double_spaces(new_value)

                        if new_value != original:
                            cell.value = new_value
                            changes_made += 1
                            change_log.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'row': row_idx,
                                'before': original[:120] + ('...' if len(original) > 120 else ''),
                                'after': new_value[:120] + ('...' if len(new_value) > 120 else ''),
                            })
                except:
                    pass

    # === Special fix for File 1: C30 speaker label ===
    if '50_analysis' in filepath:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx in range(1, ws.max_row + 1):
                try:
                    c_cell = ws.cell(row=row_idx, column=3)
                    if c_cell.value and isinstance(c_cell.value, str):
                        orig = c_cell.value
                        # Fix "Багшуваасан байна. " -> "Багш"
                        if 'Багшуваасан' in orig:
                            c_cell.value = 'Багш'
                            d_cell = ws.cell(row=row_idx, column=4)
                            if d_cell.value and isinstance(d_cell.value, str):
                                # Prepend the lost text to D column
                                d_cell.value = "Авсан байна. " + d_cell.value
                            changes_made += 1
                            change_log.append({
                                'sheet': sheet_name,
                                'cell': c_cell.coordinate,
                                'row': row_idx,
                                'before': repr(orig),
                                'after': "'Багш' (speaker label засварласан, 'Авсан байна.' текстийг D баганад нэмсэн)",
                            })
                except:
                    pass

        # Fix subject "Цлиндр" -> "Цилиндр" and extra spaces in subjects
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                d1 = ws.cell(row=1, column=4)
                if d1.value and isinstance(d1.value, str):
                    orig = d1.value
                    new_val = orig.strip()
                    if new_val != orig:
                        d1.value = new_val
                        changes_made += 1
                        change_log.append({
                            'sheet': sheet_name,
                            'cell': 'D1',
                            'row': 1,
                            'before': repr(orig),
                            'after': repr(new_val) + " (илүүдэл зай хассан)",
                        })
            except:
                pass

    wb.save(output_path)
    return changes_made, change_log


if __name__ == '__main__':
    print("=" * 70)
    print("  ФАЙЛ 1: 50_analysis (2) (1).xlsx засварлаж байна...")
    print("=" * 70)

    f1_path = r'C:\Users\Dell\Downloads\50_analysis (2) (1).xlsx'
    f1_out = r'C:\Users\Dell\Downloads\50_analysis_FIXED.xlsx'
    count1, log1 = apply_fixes(f1_path, file1_replacements, f1_out)

    print(f"\n✅ Нийт {count1} засвар хийгдлээ.")
    for item in log1:
        print(f"\n  [{item['sheet']}] Row {item['row']}, {item['cell']}:")
        print(f"    ӨМНӨ:  {item['before']}")
        print(f"    ДАРАА:  {item['after']}")

    print(f"\n📁 Хадгалсан: {f1_out}")

    print("\n" + "=" * 70)
    print("  ФАЙЛ 2: DATA_time (1) (1) (1).xlsx засварлаж байна...")
    print("=" * 70)

    f2_path = r'C:\Users\Dell\Downloads\DATA_time (1) (1) (1).xlsx'
    f2_out = r'C:\Users\Dell\Downloads\DATA_time_FIXED.xlsx'
    count2, log2 = apply_fixes(f2_path, file2_replacements, f2_out)

    print(f"\n✅ Нийт {count2} засвар хийгдлээ.")
    for item in log2:
        print(f"\n  [{item['sheet']}] Row {item['row']}, {item['cell']}:")
        print(f"    ӨМНӨ:  {item['before']}")
        print(f"    ДАРАА:  {item['after']}")

    print(f"\n📁 Хадгалсан: {f2_out}")
    print("\n" + "=" * 70)
    print(f"  🎯 НИЙТ: {count1 + count2} засвар 2 файлд хийгдлээ.")
    print("=" * 70)
