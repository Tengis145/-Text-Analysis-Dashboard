# -*- coding: utf-8 -*-
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

# Find C column errors in File 1
wb1 = openpyxl.load_workbook(r'C:\Users\Dell\Downloads\50_analysis (2) (1).xlsx')
normal_labels = {'Багш', 'Багш ', 'Сурагч', 'Сурагч ', 'Сурагчид', 'Сурагчид ', 'Сурагч Б', 'Цаг'}

for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    if ws.max_row <= 5:
        continue
    found = False
    for row_idx in range(8, ws.max_row + 1):
        try:
            c_val = ws.cell(row=row_idx, column=3).value
            if c_val is not None and isinstance(c_val, str):
                stripped = c_val.strip()
                if stripped and stripped not in {v.strip() for v in normal_labels}:
                    if not found:
                        print(f'\n=== Sheet: {sheet_name} ===')
                        found = True
                    d_val = ws.cell(row=row_idx, column=4).value
                    d_str = (str(d_val)[:80] + '...') if d_val and len(str(d_val)) > 80 else str(d_val)
                    print(f'  Row {row_idx}: C={repr(c_val)}, D={d_str}')
        except:
            pass

# Check D1 (topic/subject) for spelling
print('\n\n=== Subject/Topic cells (D1) across sheets ===')
for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    d1 = ws.cell(row=1, column=4).value
    if d1:
        print(f'  {sheet_name}: {repr(d1)}')
