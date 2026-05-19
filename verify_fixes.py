# -*- coding: utf-8 -*-
"""Засварлагдсан файлуудыг баталгаажуулах"""
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

print("=== Засварлагдсан файлууд баталгаажуулж байна ===\n")

# Verify File 1 key fixes
wb1 = openpyxl.load_workbook(r'C:\Users\Dell\Downloads\50_analysis_FIXED.xlsx')
checks = {
    'Хичээл 48': [
        (30, 3, 'C30 speaker label', lambda v: v.strip() == 'Багш'),
        (30, 4, 'D30 starts with Авсан', lambda v: v.startswith('Авсан байна.')),
        (34, 4, 'D34 no Үа', lambda v: 'Үа хариугаа' not in v),
        (34, 4, 'D34 no ёажиглаж', lambda v: 'ёажиглаж' not in v),
        (40, 4, 'D40 зүй тогтолыг', lambda v: 'зүйт огтолыг' not in v),
        (58, 4, 'D58 үржүүлээд', lambda v: 'үржшшлээд' not in v),
        (128, 4, 'D128 хүүхдүүдээ', lambda v: 'хүүдүүдээ' not in v),
        (151, 4, 'D151 гэсэн', lambda v: 'гсэн ' not in v),
        (153, 4, 'D153 баярлалаа', lambda v: 'баярллаа' not in v),
    ],
    'хичээл 45': [
        (1, 4, 'D1 Цилиндр', lambda v: v == 'Цилиндр'),
    ],
}

all_ok = True
for sheet_name, sheet_checks in checks.items():
    ws = wb1[sheet_name]
    for row, col, desc, check_fn in sheet_checks:
        cell_val = ws.cell(row=row, column=col).value
        if cell_val is None:
            print(f"  FAIL [{sheet_name}] {desc}: cell is empty!")
            all_ok = False
        elif check_fn(str(cell_val)):
            print(f"  OK   [{sheet_name}] {desc}")
        else:
            print(f"  FAIL [{sheet_name}] {desc}: {repr(str(cell_val)[:80])}")
            all_ok = False

# Verify File 2
wb2 = openpyxl.load_workbook(r'C:\Users\Dell\Downloads\DATA_time_FIXED.xlsx')
ws2 = wb2['математик']
f2_checks = [
    (69, 2, 'B69 time fix', lambda v: '26:04m' in str(v)),
    (77, 1, 'A77 no space in time', lambda v: '27:26m' in str(v) and '27: 26m' not in str(v)),
    (100, 5, 'E100 жишсэн', lambda v: 'жишсэн' in str(v)),
    (117, 5, 'E117 шидэгч', lambda v: 'шиэдгч' not in str(v)),
    (141, 5, 'E141 хялбар', lambda v: 'хялбар' in str(v)),
]

for row, col, desc, check_fn in f2_checks:
    cell_val = ws2.cell(row=row, column=col).value
    if cell_val is None:
        print(f"  FAIL [математик] {desc}: cell is empty!")
        all_ok = False
    elif check_fn(cell_val):
        print(f"  OK   [математик] {desc}")
    else:
        print(f"  FAIL [математик] {desc}: {repr(str(cell_val)[:80])}")
        all_ok = False

print()
if all_ok:
    print("Бүх засварууд амжилттай хийгдлээ!")
else:
    print("АНХААРУУЛГА: Зарим засварууд алдаатай байна.")
