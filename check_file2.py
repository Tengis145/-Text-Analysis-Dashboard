# -*- coding: utf-8 -*-
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

# File 2 full check
wb2 = openpyxl.load_workbook(r'"C:\Users\Dell\Downloads\DATA_time (1) (1).xlsx"')
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    if ws.max_row <= 5:
        continue
    print(f'=== Sheet: {sheet_name}, Rows: {ws.max_row} ===')
    # Print header
    for row_idx in range(1, 8):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    row_data.append(f'{col_letter}{row_idx}: {repr(cell.value)[:60]}')
            except:
                pass
        if row_data:
            print(' | '.join(row_data))
    print()
